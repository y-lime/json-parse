import json
from openpyxl import Workbook, load_workbook
from collections import defaultdict
import sys
import os

def make_hashable(val):
    if isinstance(val, (list, tuple)):
        return tuple(make_hashable(v) for v in val)
    elif isinstance(val, dict):
        return tuple(sorted((k, make_hashable(v)) for k, v in val.items()))
    else:
        return val

def collect_profile_keys_and_values(data):
    profile_keys = set()
    profile_values = defaultdict(set)
    def collect_keys(d, path=()):
        for k, v in d.items():
            new_path = path + (k,)
            is_dict = isinstance(v, dict)
            profile_keys.add((new_path, len(new_path), k, is_dict))
            if is_dict:
                collect_keys(v, new_path)
            else:
                try:
                    hashable_v = make_hashable(v)
                except Exception:
                    hashable_v = str(v)
                profile_values[new_path].add(hashable_v)
    for user in data:
        collect_keys(user['profile'])
    return profile_keys, profile_values

def write_header(ws, start_row, start_col, ids):
    header = ["項目名", "設定値"] + ids
    for col, val in enumerate(header, start_col):
        ws.cell(row=start_row, column=col, value=val)

def write_profile_rows(ws, start_row, start_col, profile_keys, profile_values, data, ids):
    prev_disp_name = None
    row_idx = start_row + 1
    for key_path, depth, key, is_dict in profile_keys:
        disp_name = '  ' * (depth-1) + key
        values = sorted(profile_values[key_path])
        if not values:
            row_disp_name = disp_name if prev_disp_name != disp_name else ''
            row = [row_disp_name, ''] + ['' for _ in ids]
            for col, val in enumerate(row, start_col):
                ws.cell(row=row_idx, column=col, value=val)
            prev_disp_name = disp_name
            row_idx += 1
        else:
            for value in values:
                if isinstance(value, tuple) and any(isinstance(v, tuple) for v in value):
                    disp_value = str(value)
                else:
                    disp_value = value if not isinstance(value, tuple) else str(value)
                row_disp_name = disp_name if prev_disp_name != disp_name else ''
                row = [row_disp_name, disp_value]
                for user in data:
                    v = user['profile']
                    try:
                        for part in key_path:
                            v = v[part]
                        user_hashable = make_hashable(v)
                    except (KeyError, TypeError):
                        user_hashable = None
                    if user_hashable == value:
                        row.append('◯')
                    else:
                        row.append('')
                for col, val in enumerate(row, start_col):
                    ws.cell(row=row_idx, column=col, value=val)
                prev_disp_name = disp_name
                row_idx += 1

def json_to_excel(input_path, output_path, template_path):
    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    profile_keys, profile_values = collect_profile_keys_and_values(data)
    profile_keys = sorted(profile_keys, key=lambda x: (x[1], x[3], x[0]))
    wb = load_workbook(template_path)
    ws = wb.active
    start_row = 6
    start_col = 1
    ids = [user['id'] for user in data]
    write_header(ws, start_row, start_col, ids)
    write_profile_rows(ws, start_row, start_col, profile_keys, profile_values, data, ids)
    wb.save(output_path)

def main():
    if len(sys.argv) < 2:
        print("Usage: python sript.py <input_json_file>")
        sys.exit(1)
    input_path = sys.argv[1]
    base, _ = os.path.splitext(input_path)
    output_path = base + '.xlsx'
    template_path = 'template/template.xlsx'
    json_to_excel(input_path, output_path, template_path)

if __name__ == "__main__":
    main()
